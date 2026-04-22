from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

from .database import get_connection


ITEM_TYPES = {
    "finished": "成品货",
    "semi_finished": "半成品",
    "material": "物料",
    "packaging": "包材",
}

PLATFORMS = ["拼多多", "淘宝", "京东", "抖音", "快手", "小红书", "线下"]


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_default_users(database_path: str, users: Iterable[dict[str, str]]) -> None:
    with get_connection(database_path) as conn:
        for user in users:
            existing = conn.execute("SELECT id FROM users WHERE username = ?", (user["username"],)).fetchone()
            if existing:
                continue
            ts = now()
            conn.execute(
                """
                INSERT INTO users (username, password_hash, display_name, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    user["username"],
                    generate_password_hash(user["password"]),
                    user.get("display_name") or user["username"],
                    ts,
                    ts,
                ),
            )
        _ensure_default_warehouse(conn)
        conn.commit()


def _ensure_default_warehouse(conn) -> None:
    ts = now()
    row = conn.execute("SELECT id FROM warehouses WHERE name = '默认仓库'").fetchone()
    if not row:
        conn.execute(
            "INSERT INTO warehouses (name, note, created_at, updated_at) VALUES (?, ?, ?, ?)",
            ("默认仓库", "系统默认仓库", ts, ts),
        )
        row = conn.execute("SELECT id FROM warehouses WHERE name = '默认仓库'").fetchone()
    warehouse_id = row["id"]
    conn.execute(
        "INSERT OR IGNORE INTO locations (warehouse_id, location_code, note, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
        (warehouse_id, "A-默认", "默认货位", ts, ts),
    )


def verify_user(database_path: str, username: str, password: str) -> dict | None:
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            return dict(row)
    return None


def list_users(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        return [
            dict(row)
            for row in conn.execute(
                "SELECT id, username, display_name, role, is_active, created_at, updated_at FROM users ORDER BY id"
            ).fetchall()
        ]


def save_user(database_path: str, form: dict) -> None:
    username = str(form.get("username", "")).strip()
    password = str(form.get("password", "")).strip()
    display_name = str(form.get("display_name", "")).strip() or username
    role = str(form.get("role", "staff")).strip() or "staff"
    if role not in {"admin", "boss", "warehouse", "purchase", "sales", "finance", "staff"}:
        role = "staff"
    if not username or not password:
        raise ValueError("账号和密码不能为空")
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO users (username, password_hash, display_name, role, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (username, generate_password_hash(password), display_name, role, ts, ts),
        )
        conn.commit()


def dashboard_stats(database_path: str) -> dict:
    with get_connection(database_path) as conn:
        return {
            "items": conn.execute("SELECT COUNT(*) AS c FROM items WHERE is_active = 1").fetchone()["c"],
            "finished": conn.execute("SELECT COUNT(*) AS c FROM items WHERE item_type = 'finished'").fetchone()["c"],
            "low_stock": conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM (
                    SELECT i.id, i.safety_stock, COALESCE(SUM(sm.quantity), 0) AS qty
                    FROM items i
                    LEFT JOIN stock_movements sm ON sm.item_id = i.id
                    WHERE i.is_active = 1
                    GROUP BY i.id
                ) t
                WHERE safety_stock > 0 AND qty <= safety_stock
                """
            ).fetchone()["c"],
            "receivable": conn.execute(
                "SELECT COALESCE(SUM(CASE WHEN direction='increase' THEN amount ELSE -amount END), 0) AS c FROM account_entries WHERE account_type='receivable'"
            ).fetchone()["c"],
            "payable": conn.execute(
                "SELECT COALESCE(SUM(CASE WHEN direction='increase' THEN amount ELSE -amount END), 0) AS c FROM account_entries WHERE account_type='payable'"
            ).fetchone()["c"],
            "pending_returns": conn.execute("SELECT COUNT(*) AS c FROM return_inbounds WHERE status='pending_match'").fetchone()["c"],
            "pending_orders": conn.execute("SELECT COUNT(*) AS c FROM sales_orders WHERE status IN ('pending_review', 'locked')").fetchone()["c"],
        }


def owner_dashboard(database_path: str, return_database_path: str = "") -> dict:
    current = dashboard_stats(database_path)
    quarter_start = _quarter_start()
    with get_connection(database_path) as conn:
        sales = conn.execute(
            """
            SELECT
                COALESCE(SUM(-sm.quantity), 0) AS qty,
                COALESCE(SUM(-sm.quantity * sm.unit_cost), 0) AS amount
            FROM stock_movements sm
            WHERE sm.movement_type = 'sale_out'
              AND datetime(sm.created_at) >= datetime(?)
            """,
            (quarter_start,),
        ).fetchone()
        order_stats = conn.execute(
            """
            SELECT COUNT(*) AS orders, COALESCE(SUM(total_amount), 0) AS amount
            FROM sales_orders
            WHERE datetime(created_at) >= datetime(?)
            """,
            (quarter_start,),
        ).fetchone()
        top_sales = [
            dict(row)
            for row in conn.execute(
                """
                SELECT i.item_code, i.item_name, COALESCE(SUM(-sm.quantity), 0) AS qty,
                       COALESCE(SUM(-sm.quantity * sm.unit_cost), 0) AS amount
                FROM stock_movements sm
                JOIN items i ON i.id = sm.item_id
                WHERE sm.movement_type = 'sale_out'
                  AND datetime(sm.created_at) >= datetime(?)
                GROUP BY i.id
                ORDER BY qty DESC, amount DESC
                LIMIT 10
                """,
                (quarter_start,),
            ).fetchall()
        ]
        low_stock = [
            dict(row)
            for row in conn.execute(
                """
                SELECT i.item_code, i.item_name, i.item_type, i.safety_stock, COALESCE(SUM(sm.quantity), 0) AS stock_qty
                FROM items i
                LEFT JOIN stock_movements sm ON sm.item_id = i.id
                WHERE i.is_active = 1 AND i.safety_stock > 0
                GROUP BY i.id
                HAVING stock_qty <= i.safety_stock
                ORDER BY stock_qty ASC
                LIMIT 10
                """
            ).fetchall()
        ]

    return_data = _return_dashboard(return_database_path, quarter_start)
    sales_qty = float(sales["qty"] or 0)
    return_qty = float(return_data["quarter_total"] or 0)
    return_rate = (return_qty / sales_qty * 100) if sales_qty else 0
    return {
        "stats": current,
        "quarter_start": quarter_start[:10],
        "quarter_sales_qty": sales_qty,
        "quarter_sales_amount": float(order_stats["amount"] or sales["amount"] or 0),
        "quarter_order_count": int(order_stats["orders"] or 0),
        "top_sales": top_sales,
        "low_stock": low_stock,
        "return_system": return_data,
        "return_rate": return_rate,
    }


def _return_dashboard(return_database_path: str, quarter_start: str) -> dict:
    empty = {
        "available": False,
        "quarter_registered": 0,
        "quarter_inbound": 0,
        "quarter_abnormal": 0,
        "quarter_total": 0,
        "top_returns": [],
        "recent_returns": [],
    }
    if not return_database_path:
        return empty
    try:
        conn = sqlite3.connect(return_database_path)
        conn.row_factory = sqlite3.Row
    except sqlite3.Error:
        return empty
    try:
        return {
            "available": True,
            "quarter_registered": conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM customer_returns
                WHERE datetime(imported_at) >= datetime(?)
                """,
                (quarter_start,),
            ).fetchone()["c"],
            "quarter_total": conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM scan_records
                WHERE datetime(first_scanned_at) >= datetime(?)
                """,
                (quarter_start,),
            ).fetchone()["c"],
            "quarter_inbound": conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM scan_records
                WHERE match_status = 'normal_inbound'
                  AND datetime(first_scanned_at) >= datetime(?)
                """,
                (quarter_start,),
            ).fetchone()["c"],
            "quarter_abnormal": conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM scan_records
                WHERE match_status = 'abnormal_inbound'
                  AND datetime(first_scanned_at) >= datetime(?)
                """,
                (quarter_start,),
            ).fetchone()["c"],
            "top_returns": [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT COALESCE(NULLIF(product_name, ''), '未填写商品') AS product_name,
                           COUNT(*) AS qty
                    FROM scan_records
                    WHERE datetime(first_scanned_at) >= datetime(?)
                    GROUP BY COALESCE(NULLIF(product_name, ''), '未填写商品')
                    ORDER BY qty DESC
                    LIMIT 10
                    """,
                    (quarter_start,),
                ).fetchall()
            ],
            "recent_returns": [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT barcode, product_name, customer_name, match_status, anomaly_note, first_scanned_at
                    FROM scan_records
                    ORDER BY first_scanned_at DESC
                    LIMIT 12
                    """
                ).fetchall()
            ],
        }
    except sqlite3.Error:
        return empty
    finally:
        conn.close()


def list_items(database_path: str, item_type: str = "", keyword: str = "") -> list[dict]:
    sql = """
        SELECT i.*, w.name AS warehouse_name, COALESCE(SUM(sm.quantity), 0) AS stock_qty
        FROM items i
        LEFT JOIN warehouses w ON w.id = i.default_warehouse_id
        LEFT JOIN stock_movements sm ON sm.item_id = i.id
        WHERE 1 = 1
    """
    params: list = []
    if item_type:
        sql += " AND i.item_type = ?"
        params.append(item_type)
    if keyword:
        sql += " AND (i.item_code LIKE ? OR i.item_name LIKE ? OR i.sku LIKE ? OR i.barcode LIKE ?)"
        like = f"%{keyword}%"
        params.extend([like, like, like, like])
    sql += " GROUP BY i.id ORDER BY i.updated_at DESC"
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def get_item(database_path: str, item_id: int) -> dict | None:
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        return dict(row) if row else None


def find_finished_item(database_path: str, sku: str = "", barcode: str = "", name: str = "") -> dict | None:
    with get_connection(database_path) as conn:
        for field, value in (("sku", sku), ("barcode", barcode), ("item_name", name)):
            clean = str(value or "").strip()
            if not clean:
                continue
            row = conn.execute(
                f"SELECT * FROM items WHERE item_type = 'finished' AND {field} = ? AND is_active = 1",
                (clean,),
            ).fetchone()
            if row:
                return dict(row)
    return None


def find_item_for_import(database_path: str, code_or_sku: str = "", name: str = "") -> dict | None:
    with get_connection(database_path) as conn:
        for field, value in (("item_code", code_or_sku), ("sku", code_or_sku), ("platform_sku", code_or_sku), ("barcode", code_or_sku), ("item_name", name)):
            clean = str(value or "").strip()
            if not clean:
                continue
            row = conn.execute(f"SELECT * FROM items WHERE {field} = ? AND is_active = 1", (clean,)).fetchone()
            if row:
                return dict(row)
    return None


def save_item(database_path: str, form: dict) -> dict:
    item_id = int(form.get("item_id") or 0)
    payload = {
        "item_code": str(form.get("item_code", "")).strip(),
        "item_name": str(form.get("item_name", "")).strip(),
        "item_type": str(form.get("item_type", "")).strip(),
        "category": str(form.get("category", "")).strip(),
        "unit": str(form.get("unit", "件")).strip() or "件",
        "sku": str(form.get("sku", "")).strip(),
        "barcode": str(form.get("barcode", "")).strip(),
        "spec": str(form.get("spec", "")).strip(),
        "default_warehouse_id": int(form.get("default_warehouse_id") or default_warehouse_id(database_path)),
        "default_location_id": int(form.get("default_location_id") or 0) or None,
        "platform_sku": str(form.get("platform_sku", "")).strip(),
        "supplier_id": int(form.get("supplier_id") or 0) or None,
        "lead_days": int(form.get("lead_days") or 0),
        "safety_stock": _num(form.get("safety_stock")),
        "purchase_price": _num(form.get("purchase_price")),
        "cost_price": _num(form.get("cost_price")),
        "sale_price": _num(form.get("sale_price")),
        "is_sellable": 1 if form.get("is_sellable") == "1" else 0,
        "is_producible": 1 if form.get("is_producible") == "1" else 0,
        "is_packaging": 1 if form.get("is_packaging") == "1" else 0,
        "note": str(form.get("note", "")).strip(),
        "updated_at": now(),
    }
    if not payload["item_code"] or not payload["item_name"] or payload["item_type"] not in ITEM_TYPES:
        raise ValueError("编号、名称、类型不能为空")
    with get_connection(database_path) as conn:
        if item_id:
            payload["id"] = item_id
            conn.execute(
                """
                UPDATE items SET item_code=:item_code, item_name=:item_name, item_type=:item_type,
                    category=:category, unit=:unit, sku=:sku, barcode=:barcode, spec=:spec,
                    default_warehouse_id=:default_warehouse_id, default_location_id=:default_location_id,
                    platform_sku=:platform_sku, supplier_id=:supplier_id, lead_days=:lead_days,
                    safety_stock=:safety_stock,
                    purchase_price=:purchase_price, cost_price=:cost_price, sale_price=:sale_price,
                    is_sellable=:is_sellable, is_producible=:is_producible, is_packaging=:is_packaging,
                    note=:note, updated_at=:updated_at
                WHERE id=:id
                """,
                payload,
            )
        else:
            payload["created_at"] = payload["updated_at"]
            conn.execute(
                """
                INSERT INTO items (
                    item_code, item_name, item_type, category, unit, sku, barcode, spec,
                    default_warehouse_id, default_location_id, platform_sku, supplier_id, lead_days,
                    safety_stock, purchase_price, cost_price, sale_price,
                    is_sellable, is_producible, is_packaging, note, created_at, updated_at
                ) VALUES (
                    :item_code, :item_name, :item_type, :category, :unit, :sku, :barcode, :spec,
                    :default_warehouse_id, :default_location_id, :platform_sku, :supplier_id, :lead_days,
                    :safety_stock, :purchase_price, :cost_price, :sale_price,
                    :is_sellable, :is_producible, :is_packaging, :note, :created_at, :updated_at
                )
                """,
                payload,
            )
        conn.commit()
        row = conn.execute("SELECT * FROM items WHERE item_code = ?", (payload["item_code"],)).fetchone()
        return dict(row)


def list_warehouses(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute("SELECT * FROM warehouses WHERE is_active = 1 ORDER BY name").fetchall()]


def default_warehouse_id(database_path: str) -> int:
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT id FROM warehouses ORDER BY id LIMIT 1").fetchone()
        return int(row["id"])


def save_warehouse(database_path: str, name: str, note: str = "") -> None:
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            "INSERT INTO warehouses (name, note, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (name.strip(), note.strip(), ts, ts),
        )
        conn.commit()


def list_locations(database_path: str, warehouse_id: int | None = None) -> list[dict]:
    sql = """
        SELECT l.*, w.name AS warehouse_name
        FROM locations l
        JOIN warehouses w ON w.id = l.warehouse_id
        WHERE l.is_active = 1
    """
    params = []
    if warehouse_id:
        sql += " AND l.warehouse_id = ?"
        params.append(warehouse_id)
    sql += " ORDER BY w.name, l.location_code"
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def save_location(database_path: str, form: dict) -> None:
    warehouse_id = int(form.get("warehouse_id"))
    location_code = str(form.get("location_code", "")).strip()
    if not location_code:
        raise ValueError("货位不能为空")
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO locations (warehouse_id, location_code, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (warehouse_id, location_code, str(form.get("note", "")).strip(), ts, ts),
        )
        conn.commit()


def list_partners(database_path: str, partner_type: str = "") -> list[dict]:
    sql = "SELECT * FROM partners WHERE is_active = 1"
    params = []
    if partner_type:
        sql += " AND partner_type = ?"
        params.append(partner_type)
    sql += " ORDER BY partner_type, name"
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def save_partner(database_path: str, form: dict) -> None:
    partner_type = str(form.get("partner_type", "")).strip()
    name = str(form.get("name", "")).strip()
    if partner_type not in {"customer", "supplier"} or not name:
        raise ValueError("伙伴类型和名称不能为空")
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO partners (partner_type, name, phone, contact_name, address, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                partner_type,
                name,
                str(form.get("phone", "")).strip(),
                str(form.get("contact_name", "")).strip(),
                str(form.get("address", "")).strip(),
                str(form.get("note", "")).strip(),
                ts,
                ts,
            ),
        )
        conn.commit()


def create_stock_movement(
    database_path: str,
    movement_type: str,
    item_id: int,
    warehouse_id: int,
    quantity: float,
    unit_cost: float = 0,
    source_type: str = "",
    source_no: str = "",
    document_id: int | None = None,
    partner_id: int | None = None,
    note: str = "",
    created_by: str = "",
) -> dict:
    if movement_type not in {"purchase_in", "sale_out", "return_in", "adjust_in", "adjust_out", "production_in", "consume_out"}:
        raise ValueError("无效库存动作")
    signed_qty = abs(quantity)
    if movement_type in {"sale_out", "adjust_out", "consume_out"}:
        signed_qty = -signed_qty
    ts = now()
    movement_no = f"SM{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    with get_connection(database_path) as conn:
        if signed_qty < 0:
            current_qty = _stock_quantity(conn, item_id, warehouse_id)
            if current_qty + signed_qty < -0.000001:
                raise ValueError(f"库存不足：当前库存 {current_qty:g}，本次出库 {abs(signed_qty):g}")
        conn.execute(
            """
            INSERT INTO stock_movements (
                movement_no, movement_type, item_id, warehouse_id, quantity, unit_cost,
                source_type, source_no, document_id, partner_id, note, created_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                movement_no,
                movement_type,
                item_id,
                warehouse_id,
                signed_qty,
                unit_cost,
                source_type,
                source_no,
                document_id,
                partner_id,
                note,
                created_by,
                ts,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM stock_movements WHERE movement_no = ?", (movement_no,)).fetchone()
        return dict(row)


def _stock_quantity(conn, item_id: int, warehouse_id: int) -> float:
    row = conn.execute(
        """
        SELECT COALESCE(SUM(quantity), 0) AS quantity
        FROM stock_movements
        WHERE item_id = ?
          AND warehouse_id = ?
        """,
        (item_id, warehouse_id),
    ).fetchone()
    return float(row["quantity"] or 0)


def _locked_quantity(conn, item_id: int, warehouse_id: int) -> float:
    row = conn.execute(
        """
        SELECT COALESCE(SUM(quantity), 0) AS quantity
        FROM stock_locks
        WHERE item_id = ?
          AND warehouse_id = ?
          AND status = 'locked'
        """,
        (item_id, warehouse_id),
    ).fetchone()
    return float(row["quantity"] or 0)


def _available_quantity(conn, item_id: int, warehouse_id: int) -> float:
    return _stock_quantity(conn, item_id, warehouse_id) - _locked_quantity(conn, item_id, warehouse_id)


def create_document(
    database_path: str,
    document_type: str,
    partner_id: int,
    lines: list[dict],
    source_no: str = "",
    note: str = "",
    created_by: str = "",
) -> dict:
    if document_type not in {"purchase", "sale"}:
        raise ValueError("无效单据类型")
    if not lines:
        raise ValueError("单据明细不能为空")
    ts = now()
    prefix = "PO" if document_type == "purchase" else "SO"
    document_no = source_no or f"{prefix}{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    total_amount = sum(abs(float(row["quantity"])) * float(row["unit_price"]) for row in lines)
    with get_connection(database_path) as conn:
        if document_type == "sale":
            required: dict[tuple[int, int], float] = {}
            for line in lines:
                key = (int(line["item_id"]), int(line["warehouse_id"]))
                required[key] = required.get(key, 0) + abs(float(line["quantity"]))
            for (item_id, warehouse_id), required_qty in required.items():
                current_qty = _stock_quantity(conn, item_id, warehouse_id)
                if current_qty < required_qty - 0.000001:
                    item = conn.execute("SELECT item_name FROM items WHERE id = ?", (item_id,)).fetchone()
                    item_name = item["item_name"] if item else str(item_id)
                    raise ValueError(f"{item_name} 库存不足：当前库存 {current_qty:g}，本次销售 {required_qty:g}")
        cur = conn.execute(
            """
            INSERT INTO documents (
                document_no, document_type, partner_id, total_amount, note, created_by, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (document_no, document_type, partner_id, total_amount, note, created_by, ts, ts),
        )
        document_id = cur.lastrowid
        for line in lines:
            qty = abs(float(line["quantity"]))
            unit_price = float(line["unit_price"])
            line_amount = qty * unit_price
            conn.execute(
                """
                INSERT INTO document_lines (
                    document_id, item_id, warehouse_id, quantity, unit_price, line_amount, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    document_id,
                    int(line["item_id"]),
                    int(line["warehouse_id"]),
                    qty,
                    unit_price,
                    line_amount,
                    str(line.get("note", "")),
                ),
            )
        conn.commit()

    movement_type = "purchase_in" if document_type == "purchase" else "sale_out"
    account_type = "payable" if document_type == "purchase" else "receivable"
    account_note = "采购应付增加" if document_type == "purchase" else "销售应收增加"
    for line in lines:
        create_stock_movement(
            database_path,
            movement_type,
            int(line["item_id"]),
            int(line["warehouse_id"]),
            float(line["quantity"]),
            float(line["unit_price"]),
            source_type=document_type,
            source_no=document_no,
            document_id=document_id,
            partner_id=partner_id,
            note=account_note.replace("应", ""),
            created_by=created_by,
        )
    create_account_entry(
        database_path,
        partner_id,
        account_type,
        "increase",
        total_amount,
        document_type,
        document_no,
        account_note,
        created_by,
    )
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
        return dict(row)


def create_account_entry(
    database_path: str,
    partner_id: int,
    account_type: str,
    direction: str,
    amount: float,
    source_type: str,
    source_no: str,
    note: str = "",
    created_by: str = "",
) -> None:
    ts = now()
    entry_no = f"AE{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO account_entries (
                entry_no, partner_id, account_type, direction, amount, source_type, source_no, note, created_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (entry_no, partner_id, account_type, direction, abs(amount), source_type, source_no, note, created_by, ts),
        )
        conn.commit()


def list_stock(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT i.*, w.name AS warehouse_name,
                   COALESCE(SUM(sm.quantity), 0) AS stock_qty,
                   COALESCE((
                       SELECT SUM(sl.quantity) FROM stock_locks sl
                       WHERE sl.item_id = i.id AND sl.status = 'locked'
                   ), 0) AS locked_qty
            FROM items i
            LEFT JOIN warehouses w ON w.id = i.default_warehouse_id
            LEFT JOIN stock_movements sm ON sm.item_id = i.id
            WHERE i.is_active = 1
            GROUP BY i.id
            ORDER BY i.item_type, i.item_name
            """
        ).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            item["available_qty"] = (item.get("stock_qty") or 0) - (item.get("locked_qty") or 0)
            result.append(item)
        return result


def recent_movements(database_path: str, limit: int = 100) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT sm.*, i.item_name, i.item_code, i.item_type, w.name AS warehouse_name, p.name AS partner_name
            FROM stock_movements sm
            JOIN items i ON i.id = sm.item_id
            JOIN warehouses w ON w.id = sm.warehouse_id
            LEFT JOIN partners p ON p.id = sm.partner_id
            ORDER BY sm.created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def account_summary(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT p.*, 
                COALESCE(SUM(CASE WHEN ae.direction = 'increase' THEN ae.amount ELSE -ae.amount END), 0) AS balance
            FROM partners p
            LEFT JOIN account_entries ae ON ae.partner_id = p.id
            WHERE p.is_active = 1
            GROUP BY p.id
            ORDER BY p.partner_type, p.name
            """
        ).fetchall()
        return [dict(row) for row in rows]


def list_documents(database_path: str, document_type: str = "", limit: int = 100) -> list[dict]:
    sql = """
        SELECT d.*, p.name AS partner_name, COUNT(dl.id) AS line_count
        FROM documents d
        LEFT JOIN partners p ON p.id = d.partner_id
        LEFT JOIN document_lines dl ON dl.document_id = d.id
        WHERE 1 = 1
    """
    params: list = []
    if document_type:
        sql += " AND d.document_type = ?"
        params.append(document_type)
    sql += " GROUP BY d.id ORDER BY d.created_at DESC LIMIT ?"
    params.append(limit)
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def list_account_entries(database_path: str, partner_id: int | None = None, limit: int = 200) -> list[dict]:
    sql = """
        SELECT ae.*, p.name AS partner_name, p.partner_type
        FROM account_entries ae
        JOIN partners p ON p.id = ae.partner_id
        WHERE 1 = 1
    """
    params: list = []
    if partner_id:
        sql += " AND ae.partner_id = ?"
        params.append(partner_id)
    sql += " ORDER BY ae.created_at DESC LIMIT ?"
    params.append(limit)
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def create_sales_order(database_path: str, form: dict, created_by: str = "") -> dict:
    order_no = str(form.get("order_no", "")).strip() or f"SO{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    item_ids = _form_list(form, "item_id")
    quantities = _form_list(form, "quantity")
    prices = _form_list(form, "sale_price")
    lines = []
    for index, item_id in enumerate(item_ids):
        if not item_id:
            continue
        qty = _num(quantities[index])
        if qty <= 0:
            continue
        price = _num(prices[index])
        lines.append({"item_id": int(item_id), "quantity": qty, "sale_price": price, "line_amount": qty * price})
    if not lines:
        raise ValueError("订单明细不能为空")
    total_amount = sum(row["line_amount"] for row in lines)
    ts = now()
    warehouse_id = int(form.get("warehouse_id") or default_warehouse_id(database_path))
    with get_connection(database_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO sales_orders (
                order_no, platform, shop_name, customer_id, customer_name, status, warehouse_id,
                total_amount, note, created_by, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, 'pending_review', ?, ?, ?, ?, ?, ?)
            """,
            (
                order_no,
                str(form.get("platform", "")).strip() or "拼多多",
                str(form.get("shop_name", "")).strip(),
                int(form.get("customer_id") or 0) or None,
                str(form.get("customer_name", "")).strip(),
                warehouse_id,
                total_amount,
                str(form.get("note", "")).strip(),
                created_by,
                ts,
                ts,
            ),
        )
        order_id = cur.lastrowid
        for line in lines:
            conn.execute(
                """
                INSERT INTO sales_order_lines (
                    sales_order_id, item_id, quantity, sale_price, line_amount
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (order_id, line["item_id"], line["quantity"], line["sale_price"], line["line_amount"]),
            )
        conn.commit()
        row = conn.execute("SELECT * FROM sales_orders WHERE id = ?", (order_id,)).fetchone()
        return dict(row)


def list_sales_orders(database_path: str, status: str = "", limit: int = 100) -> list[dict]:
    sql = """
        SELECT so.*, p.name AS partner_name, COUNT(sol.id) AS line_count
        FROM sales_orders so
        LEFT JOIN partners p ON p.id = so.customer_id
        LEFT JOIN sales_order_lines sol ON sol.sales_order_id = so.id
        WHERE 1 = 1
    """
    params: list = []
    if status:
        sql += " AND so.status = ?"
        params.append(status)
    sql += " GROUP BY so.id ORDER BY so.created_at DESC LIMIT ?"
    params.append(limit)
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def order_lines(database_path: str, order_id: int) -> list[dict]:
    with get_connection(database_path) as conn:
        return [
            dict(row)
            for row in conn.execute(
                """
                SELECT sol.*, i.item_code, i.item_name
                FROM sales_order_lines sol
                JOIN items i ON i.id = sol.item_id
                WHERE sol.sales_order_id = ?
                """,
                (order_id,),
            ).fetchall()
        ]


def lock_sales_order(database_path: str, order_id: int) -> None:
    ts = now()
    with get_connection(database_path) as conn:
        order = conn.execute("SELECT * FROM sales_orders WHERE id = ?", (order_id,)).fetchone()
        if not order:
            raise ValueError("订单不存在")
        if order["status"] not in {"pending_review", "locked"}:
            raise ValueError("当前状态不能锁库")
        conn.execute("DELETE FROM stock_locks WHERE source_type='sales_order' AND source_id=? AND status='locked'", (order_id,))
        lines = conn.execute("SELECT * FROM sales_order_lines WHERE sales_order_id = ?", (order_id,)).fetchall()
        required_by_item: dict[int, float] = {}
        for line in lines:
            required_by_item[line["item_id"]] = required_by_item.get(line["item_id"], 0) + float(line["quantity"])
        for item_id, required_qty in required_by_item.items():
            available_qty = _available_quantity(conn, item_id, order["warehouse_id"])
            if available_qty < required_qty - 0.000001:
                item = conn.execute("SELECT item_name FROM items WHERE id = ?", (item_id,)).fetchone()
                item_name = item["item_name"] if item else str(item_id)
                raise ValueError(f"{item_name} 可用库存不足：当前可用 {available_qty:g}，需要锁库 {required_qty:g}")
        for line in lines:
            conn.execute(
                """
                INSERT INTO stock_locks (item_id, warehouse_id, quantity, source_type, source_id, status, created_at)
                VALUES (?, ?, ?, 'sales_order', ?, 'locked', ?)
                """,
                (line["item_id"], order["warehouse_id"], line["quantity"], order_id, ts),
            )
        conn.execute("UPDATE sales_orders SET status='locked', locked_at=?, updated_at=? WHERE id=?", (ts, ts, order_id))
        conn.commit()


def ship_sales_order(database_path: str, order_id: int, logistics_company: str = "", tracking_no: str = "", created_by: str = "") -> None:
    ts = now()
    with get_connection(database_path) as conn:
        order = conn.execute("SELECT * FROM sales_orders WHERE id = ?", (order_id,)).fetchone()
        if not order:
            raise ValueError("订单不存在")
        if order["status"] not in {"pending_review", "locked"}:
            raise ValueError("当前状态不能发货")
        lines = [dict(row) for row in conn.execute("SELECT * FROM sales_order_lines WHERE sales_order_id = ?", (order_id,)).fetchall()]
    if order["status"] == "pending_review":
        lock_sales_order(database_path, order_id)
    for line in lines:
        item = get_item(database_path, line["item_id"])
        create_stock_movement(
            database_path,
            "sale_out",
            line["item_id"],
            order["warehouse_id"],
            line["quantity"],
            line["sale_price"],
            source_type="sales_order",
            source_no=order["order_no"],
            partner_id=order["customer_id"],
            note="订单发货出库",
            created_by=created_by,
        )
    if order["customer_id"]:
        create_account_entry(
            database_path,
            order["customer_id"],
            "receivable",
            "increase",
            order["total_amount"],
            "sales_order",
            order["order_no"],
            "订单应收增加",
            created_by,
        )
    with get_connection(database_path) as conn:
        conn.execute(
            "UPDATE stock_locks SET status='released', released_at=? WHERE source_type='sales_order' AND source_id=? AND status='locked'",
            (ts, order_id),
        )
        conn.execute(
            """
            UPDATE sales_orders
            SET status='shipped', shipped_at=?, logistics_company=?, tracking_no=?, updated_at=?
            WHERE id=?
            """,
            (ts, logistics_company, tracking_no, ts, order_id),
        )
        conn.commit()


def cancel_sales_order(database_path: str, order_id: int) -> None:
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            "UPDATE stock_locks SET status='released', released_at=? WHERE source_type='sales_order' AND source_id=? AND status='locked'",
            (ts, order_id),
        )
        conn.execute("UPDATE sales_orders SET status='cancelled', updated_at=? WHERE id=?", (ts, order_id))
        conn.commit()


def create_after_sale_from_return(database_path: str, return_inbound: dict, item_id: int | None) -> None:
    after_sale_no = f"AS{return_inbound['tracking_no']}"
    ts = now()
    with get_connection(database_path) as conn:
        if conn.execute("SELECT id FROM after_sales WHERE after_sale_no = ?", (after_sale_no,)).fetchone():
            return
        order = None
        if return_inbound.get("order_no"):
            order = conn.execute("SELECT * FROM sales_orders WHERE order_no = ?", (return_inbound["order_no"],)).fetchone()
        conn.execute(
            """
            INSERT INTO after_sales (
                after_sale_no, sales_order_id, tracking_no, item_id, quantity, return_quality,
                status, stock_movement_id, note, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, 'good', ?, ?, ?, ?, ?)
            """,
            (
                after_sale_no,
                order["id"] if order else None,
                return_inbound["tracking_no"],
                item_id,
                return_inbound["quantity"],
                "inbounded" if item_id else "pending_match",
                return_inbound.get("stock_movement_id"),
                "退货系统同步",
                ts,
                ts,
            ),
        )
        conn.commit()


def sync_after_sales_from_returns(database_path: str) -> int:
    before = len(list_after_sales(database_path, 100000))
    with get_connection(database_path) as conn:
        rows = [dict(row) for row in conn.execute("SELECT * FROM return_inbounds ORDER BY id").fetchall()]
    for row in rows:
        create_after_sale_from_return(database_path, row, row.get("item_id"))
    after = len(list_after_sales(database_path, 100000))
    return max(0, after - before)


def sync_return_system_records(database_path: str, return_database_path: str) -> dict:
    result = {"created": 0, "updated": 0, "skipped": 0}
    if not return_database_path or not Path(return_database_path).exists():
        return result
    source = sqlite3.connect(return_database_path)
    source.row_factory = sqlite3.Row
    try:
        rows = [
            dict(row)
            for row in source.execute(
                """
                SELECT sr.*, cr.order_no AS registered_order_no, cr.tracking_no AS registered_tracking_no,
                       cr.product_name AS registered_product_name, cr.customer_name AS registered_customer_name
                FROM scan_records sr
                LEFT JOIN customer_returns cr ON cr.id = sr.customer_return_id
                ORDER BY sr.first_scanned_at DESC
                """
            ).fetchall()
        ]
    except sqlite3.Error:
        rows = []
    finally:
        source.close()
    for row in rows:
        tracking_no = str(row.get("barcode") or "").strip()
        if not tracking_no:
            result["skipped"] += 1
            continue
        product_name = str(row.get("product_name") or row.get("registered_product_name") or "").strip()
        customer_name = str(row.get("customer_name") or row.get("registered_customer_name") or "").strip()
        order_no = str(row.get("registered_order_no") or "").strip()
        item = find_finished_item(database_path, name=product_name)
        status = "matched_inbound" if item and row.get("match_status") == "normal_inbound" else "pending_match"
        if row.get("match_status") == "abnormal_inbound":
            status = "external_abnormal"
        movement = None
        with get_connection(database_path) as conn:
            existing = conn.execute("SELECT * FROM return_inbounds WHERE tracking_no = ?", (tracking_no,)).fetchone()
        if existing:
            with get_connection(database_path) as conn:
                conn.execute(
                    """
                    UPDATE return_inbounds
                    SET order_no=?, raw_product_name=?, customer_name=?, status=?, raw_payload=?, updated_at=?
                    WHERE tracking_no=?
                    """,
                    (
                        order_no,
                        product_name,
                        customer_name,
                        status if existing["status"] == "pending_match" else existing["status"],
                        json.dumps(row, ensure_ascii=False),
                        now(),
                        tracking_no,
                    ),
                )
                conn.commit()
            result["updated"] += 1
            create_after_sale_from_return(database_path, dict(existing), existing["item_id"])
            continue
        if item and row.get("match_status") == "normal_inbound":
            movement = create_stock_movement(
                database_path,
                "return_in",
                item["id"],
                item["default_warehouse_id"] or default_warehouse_id(database_path),
                1,
                item.get("cost_price") or 0,
                source_type="return_system_db",
                source_no=tracking_no,
                note=f"退货系统数据库同步：{order_no}",
            )
        ts = str(row.get("first_scanned_at") or now())
        with get_connection(database_path) as conn:
            conn.execute(
                """
                INSERT INTO return_inbounds (
                    tracking_no, order_no, item_id, raw_product_name, raw_sku, customer_name, quantity,
                    status, stock_movement_id, raw_payload, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
                """,
                (
                    tracking_no,
                    order_no,
                    item["id"] if item else None,
                    product_name,
                    "",
                    customer_name,
                    status,
                    movement["id"] if movement else None,
                    json.dumps(row, ensure_ascii=False),
                    ts,
                    now(),
                ),
            )
            conn.commit()
            saved = dict(conn.execute("SELECT * FROM return_inbounds WHERE tracking_no = ?", (tracking_no,)).fetchone())
        create_after_sale_from_return(database_path, saved, saved.get("item_id"))
        result["created"] += 1
    return result


def list_after_sales(database_path: str, limit: int = 100) -> list[dict]:
    with get_connection(database_path) as conn:
        return [
            dict(row)
            for row in conn.execute(
                """
                SELECT af.*, i.item_name, so.order_no
                FROM after_sales af
                LEFT JOIN items i ON i.id = af.item_id
                LEFT JOIN sales_orders so ON so.id = af.sales_order_id
                ORDER BY af.created_at DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        ]


def purchase_suggestions(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT i.id, i.item_code, i.item_name, i.item_type, i.safety_stock, i.lead_days,
                   p.name AS supplier_name,
                   COALESCE(SUM(sm.quantity), 0) AS stock_qty,
                   COALESCE((
                       SELECT SUM(-sm2.quantity) FROM stock_movements sm2
                       WHERE sm2.item_id = i.id AND sm2.movement_type = 'sale_out'
                         AND datetime(sm2.created_at) >= datetime('now', '-30 days', 'localtime')
                   ), 0) AS sales_30
            FROM items i
            LEFT JOIN partners p ON p.id = i.supplier_id
            LEFT JOIN stock_movements sm ON sm.item_id = i.id
            WHERE i.is_active = 1 AND i.safety_stock > 0
            GROUP BY i.id
            HAVING stock_qty <= safety_stock OR sales_30 > 0
            ORDER BY (safety_stock - stock_qty) DESC, sales_30 DESC
            """
        ).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            avg_daily = (item["sales_30"] or 0) / 30
            target = item["safety_stock"] + avg_daily * max(item["lead_days"] or 0, 1)
            item["suggest_qty"] = max(0, round(target - (item["stock_qty"] or 0), 2))
            result.append(item)
        return result


def save_platform_settlement(database_path: str, form: dict) -> None:
    amount = _num(form.get("amount"))
    commission = _num(form.get("commission"))
    freight = _num(form.get("freight"))
    refund_amount = _num(form.get("refund_amount"))
    net_amount = amount - commission - freight - refund_amount
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO platform_settlements (
                settlement_no, platform, amount, commission, freight, refund_amount, net_amount, settled_at, note, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(form.get("settlement_no", "")).strip() or f"SET{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                str(form.get("platform", "")).strip(),
                amount,
                commission,
                freight,
                refund_amount,
                net_amount,
                str(form.get("settled_at", "")).strip(),
                str(form.get("note", "")).strip(),
                ts,
            ),
        )
        conn.commit()


def list_platform_settlements(database_path: str, limit: int = 100) -> list[dict]:
    with get_connection(database_path) as conn:
        return [dict(row) for row in conn.execute("SELECT * FROM platform_settlements ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()]


def receive_payment(database_path: str, form: dict, created_by: str = "") -> None:
    partner_id = int(form.get("partner_id"))
    amount = _num(form.get("amount"))
    entry_type = str(form.get("entry_type", "")).strip()
    source_no = str(form.get("source_no", "")).strip()
    note = str(form.get("note", "")).strip()
    if amount <= 0:
        raise ValueError("金额必须大于 0")
    if entry_type == "customer_receive":
        create_account_entry(database_path, partner_id, "receivable", "decrease", amount, "payment", source_no, note or "客户收款", created_by)
    elif entry_type == "supplier_pay":
        create_account_entry(database_path, partner_id, "payable", "decrease", amount, "payment", source_no, note or "供应商付款", created_by)
    else:
        raise ValueError("无效收付款类型")


def handle_return_inbound(database_path: str, payload: dict) -> dict:
    tracking_no = str(payload.get("tracking_no", "")).strip()
    if not tracking_no:
        raise ValueError("tracking_no 不能为空")
    qty = _num(payload.get("quantity"), 1)
    item = find_finished_item(
        database_path,
        sku=str(payload.get("sku", "")).strip(),
        barcode=str(payload.get("barcode") or payload.get("tracking_no") or "").strip(),
        name=str(payload.get("product_name", "")).strip(),
    )
    ts = now()
    status = "matched_inbound" if item else "pending_match"
    movement = None
    with get_connection(database_path) as conn:
        existing = conn.execute("SELECT * FROM return_inbounds WHERE tracking_no = ?", (tracking_no,)).fetchone()
        if existing:
            return {"status": "duplicate", "return_inbound": dict(existing)}
    if item:
        movement = create_stock_movement(
            database_path,
            "return_in",
            item["id"],
            item["default_warehouse_id"] or default_warehouse_id(database_path),
            qty,
            item.get("cost_price") or 0,
            source_type="return_system",
            source_no=tracking_no,
            note=f"退货系统入库：{payload.get('order_no', '')}",
        )
    with get_connection(database_path) as conn:
        conn.execute(
            """
            INSERT INTO return_inbounds (
                tracking_no, order_no, item_id, raw_product_name, raw_sku, customer_name, quantity,
                status, stock_movement_id, raw_payload, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                tracking_no,
                str(payload.get("order_no", "")).strip(),
                item["id"] if item else None,
                str(payload.get("product_name", "")).strip(),
                str(payload.get("sku", "")).strip(),
                str(payload.get("customer_name", "")).strip(),
                qty,
                status,
                movement["id"] if movement else None,
                json.dumps(payload, ensure_ascii=False),
                ts,
                ts,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM return_inbounds WHERE tracking_no = ?", (tracking_no,)).fetchone()
        return_inbound = dict(row)
    create_after_sale_from_return(database_path, return_inbound, item["id"] if item else None)
    return {"status": status, "return_inbound": return_inbound, "stock_movement": movement}


def list_return_inbounds(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT ri.*, i.item_code, i.item_name
            FROM return_inbounds ri
            LEFT JOIN items i ON i.id = ri.item_id
            ORDER BY ri.created_at DESC
            """
        ).fetchall()
        return [dict(row) for row in rows]


def match_return_inbound(database_path: str, return_id: int, item_id: int) -> dict:
    item = get_item(database_path, item_id)
    if not item or item["item_type"] != "finished":
        raise ValueError("只能匹配成品货")
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT * FROM return_inbounds WHERE id = ?", (return_id,)).fetchone()
        if not row:
            raise ValueError("退货记录不存在")
        ret = dict(row)
        if ret["status"] == "matched_inbound":
            raise ValueError("该退货已经入库")
    movement = create_stock_movement(
        database_path,
        "return_in",
        item["id"],
        item["default_warehouse_id"] or default_warehouse_id(database_path),
        ret["quantity"],
        item.get("cost_price") or 0,
        source_type="return_match",
        source_no=ret["tracking_no"],
        note=f"待匹配退货确认入库：{ret['order_no'] or ''}",
    )
    with get_connection(database_path) as conn:
        conn.execute(
            """
            UPDATE return_inbounds
            SET item_id = ?, status = 'matched_inbound', stock_movement_id = ?, updated_at = ?
            WHERE id = ?
            """,
            (item["id"], movement["id"], now(), return_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM return_inbounds WHERE id = ?", (return_id,)).fetchone()
        return_inbound = dict(row)
    create_after_sale_from_return(database_path, return_inbound, item["id"])
    return return_inbound


def create_bom_line(database_path: str, form: dict) -> None:
    finished_item_id = int(form.get("finished_item_id"))
    component_item_id = int(form.get("component_item_id"))
    quantity = _num(form.get("quantity"))
    if quantity <= 0:
        raise ValueError("用量必须大于 0")
    ts = now()
    with get_connection(database_path) as conn:
        conn.execute(
            "INSERT INTO bom_lines (finished_item_id, component_item_id, quantity, note, created_at) VALUES (?, ?, ?, ?, ?)",
            (finished_item_id, component_item_id, quantity, str(form.get("note", "")).strip(), ts),
        )
        conn.commit()


def list_bom_lines(database_path: str) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT bl.*, fi.item_name AS finished_name, ci.item_name AS component_name, ci.item_type AS component_type
            FROM bom_lines bl
            JOIN items fi ON fi.id = bl.finished_item_id
            JOIN items ci ON ci.id = bl.component_item_id
            ORDER BY bl.created_at DESC
            """
        ).fetchall()
        return [dict(row) for row in rows]


def create_production(database_path: str, form: dict, created_by: str = "") -> None:
    finished_item_id = int(form.get("finished_item_id"))
    warehouse_id = int(form.get("warehouse_id"))
    quantity = _num(form.get("quantity"))
    source_no = str(form.get("source_no", "")).strip() or f"PD{datetime.now().strftime('%Y%m%d%H%M%S')}"
    production_type = str(form.get("production_type", "")).strip() or "成品组装"
    production_line = str(form.get("production_line", "")).strip()
    operator_name = str(form.get("operator_name", "")).strip()
    if quantity <= 0:
        raise ValueError("生产数量必须大于 0")
    ts = now()
    with get_connection(database_path) as conn:
        if conn.execute("SELECT id FROM production_orders WHERE production_no = ?", (source_no,)).fetchone():
            raise ValueError("生产单号已存在")
        lines = [
            dict(row)
            for row in conn.execute(
                "SELECT bl.*, i.item_name AS component_name FROM bom_lines bl JOIN items i ON i.id = bl.component_item_id WHERE bl.finished_item_id = ?",
                (finished_item_id,),
            ).fetchall()
        ]
        for line in lines:
            required_qty = quantity * float(line["quantity"])
            current_qty = _stock_quantity(conn, line["component_item_id"], warehouse_id)
            if current_qty < required_qty - 0.000001:
                raise ValueError(f"{line['component_name']} 库存不足：当前库存 {current_qty:g}，需要领用 {required_qty:g}")
        conn.execute(
            """
            INSERT INTO production_orders (
                production_no, finished_item_id, warehouse_id, quantity, production_type,
                production_line, operator_name, note, created_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_no,
                finished_item_id,
                warehouse_id,
                quantity,
                production_type,
                production_line,
                operator_name,
                str(form.get("note", "")).strip(),
                created_by,
                ts,
            ),
        )
        conn.commit()
    for line in lines:
        create_stock_movement(
            database_path,
            "consume_out",
            line["component_item_id"],
            warehouse_id,
            quantity * line["quantity"],
            0,
            source_type="production",
            source_no=source_no,
            note=f"{production_type}领料 {production_line} {operator_name}".strip(),
            created_by=created_by,
        )
    item = get_item(database_path, finished_item_id)
    create_stock_movement(
        database_path,
        "production_in",
        finished_item_id,
        warehouse_id,
        quantity,
        item.get("cost_price") or 0 if item else 0,
        source_type="production",
        source_no=source_no,
        note=f"{production_type}入库 {production_line} {operator_name}".strip(),
        created_by=created_by,
    )


def list_production_orders(database_path: str, limit: int = 100) -> list[dict]:
    with get_connection(database_path) as conn:
        rows = conn.execute(
            """
            SELECT po.*, i.item_code, i.item_name, w.name AS warehouse_name
            FROM production_orders po
            JOIN items i ON i.id = po.finished_item_id
            JOIN warehouses w ON w.id = po.warehouse_id
            ORDER BY po.created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def create_import_template(export_dir: str, kind: str) -> str:
    templates = {
        "items": ("商品导入模板.xlsx", ["类型", "编号", "名称", "SKU", "平台SKU", "条码", "规格", "单位", "供应商", "安全库存", "采购价", "成本价", "销售价", "采购提前期"]),
        "orders": ("订单导入模板.xlsx", ["平台", "店铺", "订单号", "商品编号或SKU", "商品名称", "数量", "销售单价", "客户姓名", "物流公司", "快递单号", "备注"]),
        "purchase": ("采购导入模板.xlsx", ["供应商", "采购单号", "商品编号或SKU", "商品名称", "数量", "采购单价", "备注"]),
        "stock": ("库存调整导入模板.xlsx", ["动作", "商品编号或SKU", "商品名称", "数量", "单价或成本", "来源单号", "备注"]),
    }
    if kind not in templates:
        raise ValueError("未知模板类型")
    filename, headers = templates[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(headers)
    if kind == "orders":
        ws.append(["拼多多", "默认店铺", "PDD20260001", "CP001", "测试成品", 1, 99, "张三", "顺丰", "SF10001", ""])
    elif kind == "items":
        ws.append(["成品货", "CP001", "测试成品", "SKU001", "PDD-SKU001", "", "默认规格", "件", "默认供应商", 10, 50, 60, 99, 3])
    elif kind == "purchase":
        ws.append(["默认供应商", "PO20260001", "CP001", "测试成品", 10, 50, ""])
    else:
        ws.append(["盘盈入库", "CP001", "测试成品", 1, 60, "ADJ20260001", ""])
    Path(export_dir).mkdir(parents=True, exist_ok=True)
    path = Path(export_dir) / filename
    wb.save(path)
    return str(path)


def import_excel(database_path: str, file_path: str, kind: str, created_by: str = "") -> dict:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = {"created": 0, "skipped": 0, "errors": []}
    for idx, row in enumerate(rows, start=2):
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        try:
            if kind == "orders":
                _import_order_row(database_path, row, created_by)
            elif kind == "items":
                _import_item_row(database_path, row)
            elif kind == "purchase":
                _import_purchase_row(database_path, row, created_by)
            elif kind == "stock":
                _import_stock_row(database_path, row, created_by)
            else:
                raise ValueError("未知导入类型")
            result["created"] += 1
        except Exception as exc:
            result["skipped"] += 1
            result["errors"].append(f"第{idx}行：{exc}")
    return result


def _import_order_row(database_path: str, row: tuple, created_by: str) -> None:
    platform, shop, order_no, sku, item_name, qty, price, customer_name, logistics, tracking, note = (list(row) + [""] * 11)[:11]
    if not order_no:
        raise ValueError("订单号不能为空")
    with get_connection(database_path) as conn:
        if conn.execute("SELECT id FROM sales_orders WHERE order_no = ?", (str(order_no).strip(),)).fetchone():
            raise ValueError("订单号已存在")
    item = find_item_for_import(database_path, str(sku or ""), str(item_name or ""))
    if not item:
        raise ValueError("找不到商品，请先导入商品资料")
    create_sales_order(
        database_path,
        {
            "platform": platform or "拼多多",
            "shop_name": shop or "默认店铺",
            "order_no": str(order_no).strip(),
            "customer_name": customer_name or "",
            "warehouse_id": item["default_warehouse_id"] or default_warehouse_id(database_path),
            "item_id": [item["id"]],
            "quantity": [qty or 1],
            "sale_price": [price or item.get("sale_price") or 0],
            "note": note or "",
        },
        created_by,
    )
    if logistics or tracking:
        with get_connection(database_path) as conn:
            conn.execute(
                "UPDATE sales_orders SET logistics_company=?, tracking_no=?, updated_at=? WHERE order_no=?",
                (str(logistics or ""), str(tracking or ""), now(), str(order_no).strip()),
            )
            conn.commit()


def _import_item_row(database_path: str, row: tuple) -> None:
    type_name, code, name, sku, platform_sku, barcode, spec, unit, supplier_name, safety, purchase, cost, sale, lead_days = (list(row) + [""] * 14)[:14]
    type_map = {v: k for k, v in ITEM_TYPES.items()}
    item_type = type_map.get(str(type_name).strip(), str(type_name or "").strip())
    supplier_id = None
    if supplier_name:
        supplier_id = _ensure_partner(database_path, "supplier", str(supplier_name).strip())
    save_item(
        database_path,
        {
            "item_type": item_type,
            "item_code": code,
            "item_name": name,
            "sku": sku,
            "platform_sku": platform_sku,
            "barcode": barcode,
            "spec": spec,
            "unit": unit or "件",
            "supplier_id": supplier_id or "",
            "safety_stock": safety or 0,
            "purchase_price": purchase or 0,
            "cost_price": cost or 0,
            "sale_price": sale or 0,
            "lead_days": lead_days or 0,
            "is_sellable": "1" if item_type == "finished" else "0",
        },
    )


def _import_purchase_row(database_path: str, row: tuple, created_by: str) -> None:
    supplier_name, source_no, sku, item_name, qty, cost, note = (list(row) + [""] * 7)[:7]
    supplier_id = _ensure_partner(database_path, "supplier", str(supplier_name or "默认供应商").strip())
    item = find_item_for_import(database_path, str(sku or ""), str(item_name or ""))
    if not item:
        raise ValueError("找不到商品")
    create_document(
        database_path,
        "purchase",
        supplier_id,
        [{"item_id": item["id"], "warehouse_id": item["default_warehouse_id"] or default_warehouse_id(database_path), "quantity": qty or 1, "unit_price": cost or item.get("purchase_price") or 0}],
        str(source_no or "").strip(),
        str(note or "").strip(),
        created_by,
    )


def _import_stock_row(database_path: str, row: tuple, created_by: str) -> None:
    action, sku, item_name, qty, cost, source_no, note = (list(row) + [""] * 7)[:7]
    action_map = {"盘盈入库": "adjust_in", "盘亏出库": "adjust_out", "销售出库": "sale_out", "领用出库": "consume_out"}
    movement_type = action_map.get(str(action).strip(), str(action or "adjust_in").strip())
    item = find_item_for_import(database_path, str(sku or ""), str(item_name or ""))
    if not item:
        raise ValueError("找不到商品")
    create_stock_movement(
        database_path,
        movement_type,
        item["id"],
        item["default_warehouse_id"] or default_warehouse_id(database_path),
        qty or 1,
        cost or item.get("cost_price") or 0,
        source_type="import",
        source_no=str(source_no or "").strip(),
        note=str(note or "").strip(),
        created_by=created_by,
    )


def export_report(database_path: str, export_dir: str, kind: str) -> str:
    exporters = {
        "stock": ("库存汇总.xlsx", ["类型", "编号", "名称", "仓库", "库存", "锁定", "可用", "安全库存"], list_stock),
        "accounts": ("账目汇总.xlsx", ["类型", "名称", "余额"], account_summary),
        "returns": ("退货对接.xlsx", ["时间", "状态", "快递单号", "订单号", "原始商品", "SKU", "匹配商品", "数量"], list_return_inbounds),
        "purchase": ("采购汇总.xlsx", ["时间", "单号", "供应商", "金额", "状态"], lambda db: list_documents(db, "purchase", 1000)),
        "sales": ("销售汇总.xlsx", ["时间", "单号", "客户", "金额", "状态"], lambda db: list_documents(db, "sale", 1000)),
        "orders": ("订单汇总.xlsx", ["时间", "订单号", "平台", "店铺", "客户", "金额", "状态", "物流", "快递单号"], lambda db: list_sales_orders(db, "", 1000)),
        "settlements": ("平台对账.xlsx", ["时间", "对账单号", "平台", "成交额", "佣金", "运费", "退款", "净额"], lambda db: list_platform_settlements(db, 1000)),
    }
    if kind not in exporters:
        raise ValueError("未知导出类型")
    filename, headers, loader = exporters[kind]
    rows = loader(database_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(headers)
    for row in rows:
        if kind == "stock":
            ws.append([ITEM_TYPES.get(row["item_type"], row["item_type"]), row["item_code"], row["item_name"], row["warehouse_name"], row["stock_qty"], row["locked_qty"], row["available_qty"], row["safety_stock"]])
        elif kind == "accounts":
            ws.append([{"customer": "客户", "supplier": "供应商"}.get(row["partner_type"], row["partner_type"]), row["name"], row["balance"]])
        elif kind == "returns":
            ws.append([row["created_at"], row["status"], row["tracking_no"], row["order_no"], row["raw_product_name"], row["raw_sku"], row.get("item_name", ""), row["quantity"]])
        elif kind == "orders":
            ws.append([row["created_at"], row["order_no"], row["platform"], row["shop_name"], row.get("partner_name") or row["customer_name"], row["total_amount"], row["status"], row["logistics_company"], row["tracking_no"]])
        elif kind == "settlements":
            ws.append([row["created_at"], row["settlement_no"], row["platform"], row["amount"], row["commission"], row["freight"], row["refund_amount"], row["net_amount"]])
        else:
            ws.append([row["created_at"], row["document_no"], row.get("partner_name", ""), row["total_amount"], row["status"]])
    Path(export_dir).mkdir(parents=True, exist_ok=True)
    path = Path(export_dir) / filename
    wb.save(path)
    return str(path)


def _num(value, default: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _quarter_start() -> str:
    today = datetime.now()
    start_month = ((today.month - 1) // 3) * 3 + 1
    return datetime(today.year, start_month, 1).strftime("%Y-%m-%d 00:00:00")


def _form_list(form: dict, key: str) -> list:
    if hasattr(form, "getlist"):
        return form.getlist(key)
    value = form.get(key)
    return value if isinstance(value, list) else [value]


def _ensure_partner(database_path: str, partner_type: str, name: str) -> int:
    if not name:
        name = "默认供应商" if partner_type == "supplier" else "默认客户"
    with get_connection(database_path) as conn:
        row = conn.execute("SELECT id FROM partners WHERE partner_type=? AND name=?", (partner_type, name)).fetchone()
        if row:
            return int(row["id"])
        ts = now()
        cur = conn.execute(
            "INSERT INTO partners (partner_type, name, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (partner_type, name, ts, ts),
        )
        conn.commit()
        return int(cur.lastrowid)
